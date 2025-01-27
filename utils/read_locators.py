import openpyxl


def read_locators(locator_file="locator.xlsx"):
    """Reads locators from an Excel file and returns them as a dictionary."""
    try:
        workbook = openpyxl.load_workbook(locator_file)
        sheet = workbook.active

        locators = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            element_name = row[0]  # Element name
            locator_type = row[1]  # Locator type (XPATH)
            locator_value = row[2]  # Locator value
            if locator_type.upper() == "XPATH":
                locators[element_name] = locator_value
            elif locator_type.upper() == "ID":
                locators[element_name] = locator_value
            elif locator_type.upper() == "CLASS":
                locators[element_name] = locator_value

        return locators

    except Exception as e:
        print(f"Error reading locators: {e}")
        return {}


def read_msg(locator_file="message.xlsx"):
    try:
        workbook = openpyxl.load_workbook(locator_file)
        sheet = workbook.active

        # Assuming the first row contains headers, and data starts from the second row
        headers = [cell.value for cell in sheet[1]]  # Read header row
        data = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            for key, value in zip(headers, row):
                data[key] = value  # Map headers to values
        return data
    except Exception as e:
        print(f"Error reading message from Excel: {str(e)}")
        return {}
